from datetime import datetime

import os
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment


class ClassificationExcel:

    def __init__(self, order_xlsx_filename, partner_info_xlsx_filename, path='result'):
        # 주문목록
        df = pd.read_excel(order_xlsx_filename)
        df = df.rename(columns=df.iloc[1])
        df = df.drop([df.index[0], df.index[1]])
        df = df.reset_index(drop=True)
        self.order_list = df
        self.path = path
        os.makedirs(path, exist_ok=True)

        # 파트너목록
        df_partners = pd.read_excel(partner_info_xlsx_filename)

        self.brands = df_partners['브랜드'].to_list()
        self.partners = df_partners['업체명'].to_list()


    def classify(self):

        for i, row in self.order_list.iterrows():
            partner_name = ''
            brand_name = ''

            print(f'{row["상품명"]} 을 분류합니다.')
            for j in range(len(self.brands)):
                if self.brands[j] in row['상품명']:
                    brand_name = self.brands[j]
                    partner_name = self.partners[j]
                    break #브랜드 못찾으면 빈문자열출력->직접excel에 입력
            print(f'브랜드:{brand_name}')
            print(f'업체명:{partner_name}')

            if brand_name != '': #브랜드를 찾음
                df_filtered = self.order_list[self.order_list['상품명'].str.contains(brand_name)]
                df_filtered.to_excel(os.path.join(self.path, f'[스쿠몰] {partner_name}.xlsx'))
            else:
                print('브랜드를 못찾았습니다.', row['상품명'])

    def set_count(self):
        file_name = '20221113/[스쿠몰] 다온마켓.xlsx'
        wb = load_workbook(file_name) #엑셀파일열기
        ws = wb.active
        print('value:', ws['B1'].value)
        print('value:', ws['B2'].value)

        # 개수 세기
        row_cnt = ws.max_row - 1
        print('cnt:', row_cnt)

        # 열 삽입
        ws.insert_rows(1)
        ws.insert_rows(1)

        now_day = datetime.now().strftime('%Y-%m-%d')

        # A1
        ws['A1'] = f'발송요청내역 [총 {row_cnt}건] {now_day}'
        ws['A1'].font = Font(size=11, bold=True)
        ws.merge_cells('A1:U1')
        ws['A1'].alignment = Alignment(horizontal='left')

        wb.save(file_name)





if __name__ == '__main__':
    ce = ClassificationExcel('주문목록20221112.xlsx', '파트너목록.xlsx', '20221113')
    ce.classify()
    ce.set_count()
