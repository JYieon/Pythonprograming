#이메일을 여러사람에게 전송
import os
import smtplib
from email.mime.text import MIMEText

from openpyxl.reader.excel import load_workbook



class EmailSender:
    email_addr = None
    password = None
    smtp_server_map = {
        'gmail.com':'smtp.gmail.com',
        'naver.com':'smtp.naver.com'
    }
    smtp_server = None

    def __init__(self, email_addr, password):
        self.email_addr = email_addr
        self.password = password
        self.smtp_server = self.smtp_server_map[email_addr.split('@')[1]]
        print(self.smtp_server)

    def send_email(self, msg, from_addr, to_addr):
        """
        :param msg: 보낼 메세지
        :param from_addr: 보내는 사람
        :param to_addr: 받는 사람
        :return:
        """
        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            msg = MIMEText(msg)
            msg['From'] = from_addr
            msg['To'] = to_addr
            msg['Subject'] = "이메일 전송 테스트"
            print(msg.as_string())

            smtp.starttls()
            smtp.login(self.email_addr, self.password)
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.as_string())
            smtp.quit()
        print('이메일 전송이 완료 되었습니다.')

    def send_all_emails(self, filename):
        print(f'{filename}에 있는 이메일과 내용을 이용해 메일을 보냅니다.')
        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows():
            print(row[0], row[1].value)

if __name__ == '__main__':
    gmail_address = GOOGLE_ADD
    gmail_password = GOOGLE_PW
    naver_address = NAVER_ADD
    naver_password = NAVER_PW
    es_gmail = EmailSender(gmail_address, gmail_password)
    es_naver = EmailSender(naver_address, naver_password)
    # es_gmail.send_email('테스트 입니다.\n구글 이메일에서 보냄', from_addr=gmail_address, to_addr=naver_address)
    # es_naver.send_email('테스트 입니다.\n네이버 이메일에서 보냄', from_addr=naver_address, to_addr=gmail_address)
    es_gmail.send_all_emails('이메일리스트.xlsx')
    es_naver.send_all_emails('이메일리스트.xlsx')
